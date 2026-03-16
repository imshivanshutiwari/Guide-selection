"""
Guide Selection Project — Main Flask Application.
Full-stack web app for student-guide allocation.
"""

import os
import io
from datetime import datetime, timezone
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, session, send_file)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
import bcrypt
from werkzeug.utils import secure_filename
import PyPDF2
from scholarly import scholarly
from flask_mail import Mail, Message

from models import db, User, Student, Guide, Preference, Allocation, Notification, AuditLog
from matching import run_matching

# ─── App Configuration ──────────────────────────────────────────
app = Flask(__name__)
app.config['SECRET_KEY'] = 'guide-selection-secret-key-2026'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///guide_selection.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

UPLOAD_FOLDER = os.path.join('static', 'uploads', 'sops')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max limit

# Mail Config
app.config['MAIL_SERVER'] = 'localhost'
app.config['MAIL_PORT'] = 1025
app.config['MAIL_DEFAULT_SENDER'] = 'noreply@college.edu'
app.config['MAIL_SUPPRESS_SEND'] = True
mail = Mail(app)

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def send_local_email(to, subject, body):
    """Simulate sending an email by logging to a local file."""
    os.makedirs('emails', exist_ok=True)
    filename = f"emails/{datetime.now().strftime('%Y%m%d_%H%M%S')}_{to.replace('@', '_')}.txt"
    try:
        with open(filename, 'w') as f:
            f.write(f"To: {to}\nSubject: {subject}\n\n{body}")
    except: pass


# ─── Role Decorators ────────────────────────────────────────────
def role_required(*roles):
    """Decorator to restrict access to specific roles."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role not in roles:
                flash('Access denied. Insufficient permissions.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def log_action(action, target='', details=''):
    """Helper to create an audit log entry."""
    audit = AuditLog(
        actor_id=current_user.id if current_user.is_authenticated else None,
        action=action,
        target=target,
        details=details
    )
    db.session.add(audit)
    db.session.commit()


# ─── Auth Routes ─────────────────────────────────────────────────
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        user = User.query.filter_by(email=email).first()
        if user and bcrypt.checkpw(password.encode('utf-8'), user.password_hash.encode('utf-8')):
            login_user(user)
            log_action('login', target=user.email)
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password.', 'error')

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'student')
        department = request.form.get('department', 'Computer Science')

        if User.query.filter_by(email=email).first():
            flash('Email already registered.', 'error')
            return render_template('register.html')

        if len(password) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return render_template('register.html')

        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        colors = ['#6C63FF', '#FF6584', '#43E97B', '#F7971E', '#00C9FF', '#FC5C7D']
        import random
        user = User(
            name=name, email=email, password_hash=hashed,
            role=role, department=department,
            avatar_color=random.choice(colors)
        )
        db.session.add(user)
        db.session.flush()

        if role == 'student':
            cgpa = float(request.form.get('cgpa', 0))
            interests = request.form.getlist('interests')
            
            # Handle SOP PDF Upload
            sop_file = request.files.get('sop_file')
            sop_path = ''
            sop_score = 0.0
            
            if sop_file and sop_file.filename.endswith('.pdf'):
                filename = secure_filename(f"{user.id}_{sop_file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                sop_file.save(filepath)
                sop_path = f"uploads/sops/{filename}"
                
                # Basic NLP Score
                try:
                    with open(filepath, 'rb') as f:
                        pdf = PyPDF2.PdfReader(f)
                        text = " ".join([page.extract_text() for page in pdf.pages if page.extract_text()]).lower()
                        for interest in interests:
                            if interest.lower() in text:
                                sop_score += 5.0
                except Exception as e:
                    print("PDF Error:", e)

            student = Student(user_id=user.id, cgpa=cgpa, area_of_interest=interests, sop_url=sop_path, sop_score=sop_score)
            db.session.add(student)
        elif role == 'guide':
            areas = request.form.getlist('research_areas')
            capacity = int(request.form.get('capacity', 5))
            guide = Guide(user_id=user.id, research_areas=areas, capacity=capacity)
            db.session.add(guide)

        # Welcome notification
        notif = Notification(
            user_id=user.id, type='success', title='Welcome!',
            message=f'Welcome to the Guide Selection System, {name}!'
        )
        db.session.add(notif)
        db.session.commit()
        log_action('register', target=email, details=f'role={role}')
        
        send_local_email(email, "Welcome to GuideSelect!", f"Hello {name},\nWelcome to the platform!")

        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    log_action('logout', target=current_user.email)
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ─── Dashboard Router ───────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'student':
        return redirect(url_for('student_dashboard'))
    elif current_user.role == 'guide':
        return redirect(url_for('guide_dashboard'))
    elif current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('login'))


# ─── Student Routes ─────────────────────────────────────────────
@app.route('/student/dashboard')
@login_required
@role_required('student')
def student_dashboard():
    student = current_user.student_profile
    allocation = Allocation.query.filter_by(student_id=student.id).first() if student else None
    guide = db.session.get(Guide, allocation.guide_id) if allocation else None
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(10).all()
    guides = Guide.query.all()

    return render_template('student/dashboard.html',
                         student=student, allocation=allocation,
                         assigned_guide=guide, notifications=notifications,
                         guides=guides)


@app.route('/student/preferences', methods=['GET', 'POST'])
@login_required
@role_required('student')
def student_preferences():
    student = current_user.student_profile
    guides = Guide.query.all()
    existing_pref = Preference.query.filter_by(student_id=student.id).first()

    if request.method == 'POST':
        c1 = request.form.get('choice_1')
        c2 = request.form.get('choice_2')
        c3 = request.form.get('choice_3')

        # Validate no duplicates
        choices = [c for c in [c1, c2, c3] if c]
        if len(choices) != len(set(choices)):
            flash('Each preference must be a different guide.', 'error')
            return render_template('student/preferences.html',
                                 student=student, guides=guides, preference=existing_pref)

        if existing_pref:
            existing_pref.choice_1_id = int(c1) if c1 else None
            existing_pref.choice_2_id = int(c2) if c2 else None
            existing_pref.choice_3_id = int(c3) if c3 else None
            existing_pref.submitted_at = datetime.now(timezone.utc)
        else:
            pref = Preference(
                student_id=student.id,
                choice_1_id=int(c1) if c1 else None,
                choice_2_id=int(c2) if c2 else None,
                choice_3_id=int(c3) if c3 else None
            )
            db.session.add(pref)

        notif = Notification(
            user_id=current_user.id, type='success', title='Preferences Saved',
            message='Your guide preferences have been submitted successfully.'
        )
        db.session.add(notif)
        db.session.commit()
        log_action('submit_preferences', target=str(student.id), details=f'choices={choices}')
        flash('Preferences submitted successfully!', 'success')
        return redirect(url_for('student_dashboard'))

    return render_template('student/preferences.html',
                         student=student, guides=guides, preference=existing_pref)


# ─── Guide Routes ───────────────────────────────────────────────
@app.route('/guide/dashboard')
@login_required
@role_required('guide')
def guide_dashboard():
    guide = current_user.guide_profile
    allocations = Allocation.query.filter_by(guide_id=guide.id).all() if guide else []
    # Get students who listed this guide in their preferences
    applicants = []
    prefs = Preference.query.filter(
        (Preference.choice_1_id == guide.id) |
        (Preference.choice_2_id == guide.id) |
        (Preference.choice_3_id == guide.id)
    ).all()
    for pref in prefs:
        student = db.session.get(Student, pref.student_id)
        if student:
            rank = 0
            if pref.choice_1_id == guide.id:
                rank = 1
            elif pref.choice_2_id == guide.id:
                rank = 2
            elif pref.choice_3_id == guide.id:
                rank = 3
            applicants.append({
                'student': student,
                'rank': rank,
                'score': guide.applicant_score(student),
                'allocation': Allocation.query.filter_by(student_id=student.id, guide_id=guide.id).first()
            })
    applicants.sort(key=lambda x: x['score'], reverse=True)

    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(10).all()

    return render_template('guide/dashboard.html',
                         guide=guide, allocations=allocations,
                         applicants=applicants, notifications=notifications)


@app.route('/guide/profile', methods=['GET', 'POST'])
@login_required
@role_required('guide')
def guide_profile():
    guide = current_user.guide_profile

    if request.method == 'POST':
        guide.bio = request.form.get('bio', '')
        guide.capacity = int(request.form.get('capacity', 5))
        guide.designation = request.form.get('designation', 'Assistant Professor')
        areas = request.form.get('research_areas_text', '')
        guide.research_areas = [a.strip() for a in areas.split(',') if a.strip()]
        db.session.commit()
        log_action('update_profile', target=str(guide.id))
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('guide_dashboard'))

    return render_template('guide/profile.html', guide=guide)


@app.route('/guide/fetch-scholar', methods=['POST'])
@login_required
@role_required('guide')
def fetch_scholar():
    guide = current_user.guide_profile
    try:
        search_query = scholarly.search_author(current_user.name)
        author = next(search_query)
        scholarly.fill(author, sections=['publications'])
        
        pubs = []
        for pub in author['publications'][:5]:
            pubs.append(pub['bib'].get('title', 'Unknown Publication'))
            
        guide.scholar_id = author['scholar_id']
        guide.publications = pubs
        db.session.commit()
        log_action('fetch_scholar', target=str(guide.id))
        flash('Successfully fetched publications from Google Scholar!', 'success')
    except StopIteration:
        flash('Could not find profile on Google Scholar with your exact name.', 'error')
    except Exception as e:
        flash(f'Google Scholar API Error: {str(e)}', 'error')
        
    return redirect(url_for('guide_profile'))


@app.route('/guide/respond', methods=['POST'])
@login_required
@role_required('guide')
def guide_respond():
    guide = current_user.guide_profile
    student_id = int(request.form.get('student_id'))
    action = request.form.get('action')  # accept, reject, waitlist

    alloc = Allocation.query.filter_by(student_id=student_id, guide_id=guide.id).first()
    student = db.session.get(Student, student_id)

    if action == 'accept' and alloc:
        alloc.status = 'confirmed'
        notif = Notification(
            user_id=student.user_id, type='success', title='Guide Confirmed!',
            message=f'{current_user.name} has confirmed your allocation.'
        )
        db.session.add(notif)
    elif action == 'reject' and alloc:
        alloc.status = 'rejected'
        guide.current_load = max(0, guide.current_load - 1)
        notif = Notification(
            user_id=student.user_id, type='error', title='Allocation Rejected',
            message=f'{current_user.name} has rejected your allocation. Admin will reassign.'
        )
        db.session.add(notif)
    elif action == 'waitlist' and alloc:
        alloc.status = 'waitlisted'
        notif = Notification(
            user_id=student.user_id, type='warning', title='Waitlisted',
            message=f'You have been waitlisted by {current_user.name}.'
        )
        db.session.add(notif)

    db.session.commit()
    log_action('guide_response', target=str(student_id), details=f'action={action}')
    flash(f'Student {action}ed successfully.', 'success')
    return redirect(url_for('guide_dashboard'))


# ─── Admin Routes ────────────────────────────────────────────────
@app.route('/admin/dashboard')
@login_required
@role_required('admin')
def admin_dashboard():
    students = Student.query.all()
    guides = Guide.query.all()
    allocations = Allocation.query.all()
    unmatched = [s for s in students if not Allocation.query.filter_by(student_id=s.id).first()]
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(10).all()
    audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(20).all()

    total_capacity = sum(g.capacity for g in guides)
    total_allocated = sum(1 for a in allocations if a.status in ('allocated', 'confirmed'))

    stats = {
        'total_students': len(students),
        'total_guides': len(guides),
        'total_allocated': total_allocated,
        'total_unmatched': len(unmatched),
        'total_capacity': total_capacity,
        'completion_pct': round((total_allocated / len(students) * 100) if students else 0, 1),
        'pref_submitted': Preference.query.count(),
    }

    # Guide load data for charts
    guide_load_data = []
    for g in guides:
        guide_load_data.append({
            'name': g.user.name,
            'capacity': g.capacity,
            'load': g.current_load,
            'areas': ', '.join(g.research_areas[:3])
        })

    # Preference distribution
    pref_dist = {'choice_1': 0, 'choice_2': 0, 'choice_3': 0, 'fallback': 0}
    for a in allocations:
        if a.preference_rank == 1:
            pref_dist['choice_1'] += 1
        elif a.preference_rank == 2:
            pref_dist['choice_2'] += 1
        elif a.preference_rank == 3:
            pref_dist['choice_3'] += 1
        else:
            pref_dist['fallback'] += 1

    return render_template('admin/dashboard.html',
                         students=students, guides=guides,
                         allocations=allocations, unmatched=unmatched,
                         stats=stats, guide_load_data=guide_load_data,
                         pref_dist=pref_dist, notifications=notifications,
                         audit_logs=audit_logs)


@app.route('/admin/run-matching', methods=['POST'])
@login_required
@role_required('admin')
def admin_run_matching():
    result = run_matching(db, Student, Guide, Preference, Allocation, Notification, AuditLog)
    flash(f'Matching complete! Phase 1: {result["phase1_matched"]}, Phase 2: {result["phase2_matched"]}, '
          f'Unmatched: {result["unmatched"]}', 'success')
    session['matching_result'] = result
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/override', methods=['POST'])
@login_required
@role_required('admin')
def admin_override():
    student_id = int(request.form.get('student_id'))
    guide_id = int(request.form.get('guide_id'))

    # Remove existing allocation if any
    existing = Allocation.query.filter_by(student_id=student_id).first()
    if existing:
        old_guide = db.session.get(Guide, existing.guide_id)
        if old_guide:
            old_guide.current_load = max(0, old_guide.current_load - 1)
        db.session.delete(existing)

    # Create new allocation
    guide = db.session.get(Guide, guide_id)
    student = db.session.get(Student, student_id)
    alloc = Allocation(
        student_id=student_id,
        guide_id=guide_id,
        status='allocated',
        method='manual',
        preference_rank=0,
        allocated_at=datetime.now(timezone.utc)
    )
    db.session.add(alloc)
    guide.current_load += 1

    # Notify
    notif = Notification(
        user_id=student.user_id, type='info', title='Guide Reassigned',
        message=f'You have been manually assigned to {guide.user.name} by admin.'
    )
    db.session.add(notif)
    db.session.commit()
    log_action('admin_override', target=f'student={student_id}', details=f'assigned to guide={guide_id}')
    flash(f'Manually assigned {student.user.name} → {guide.user.name}', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/analytics')
@login_required
@role_required('admin')
def admin_analytics():
    guides = Guide.query.all()
    allocations = Allocation.query.all()
    students = Student.query.all()

    # Most preferred guides
    guide_demand = {}
    for g in guides:
        count = Preference.query.filter(
            (Preference.choice_1_id == g.id) |
            (Preference.choice_2_id == g.id) |
            (Preference.choice_3_id == g.id)
        ).count()
        guide_demand[g.user.name] = count

    # Method distribution
    method_dist = {}
    for a in allocations:
        method_dist[a.method] = method_dist.get(a.method, 0) + 1

    # CGPA distribution of allocated students
    cgpa_ranges = {'Below 7': 0, '7-8': 0, '8-9': 0, '9+': 0}
    for s in students:
        if s.cgpa < 7:
            cgpa_ranges['Below 7'] += 1
        elif s.cgpa < 8:
            cgpa_ranges['7-8'] += 1
        elif s.cgpa < 9:
            cgpa_ranges['8-9'] += 1
        else:
            cgpa_ranges['9+'] += 1

    return render_template('admin/analytics.html',
                         guides=guides, guide_demand=guide_demand,
                         method_dist=method_dist, cgpa_ranges=cgpa_ranges,
                         allocations=allocations, students=students)


@app.route('/admin/export')
@login_required
@role_required('admin')
def admin_export():
    """Export allocations to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Allocations"

        # Header styling
        header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['S.No', 'Student Name', 'Email', 'CGPA', 'Department',
                   'Guide Name', 'Status', 'Method', 'Preference Rank', 'Allocated At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        allocations = Allocation.query.all()
        for row_idx, alloc in enumerate(allocations, 2):
            student = db.session.get(Student, alloc.student_id)
            guide = db.session.get(Guide, alloc.guide_id)
            data = [
                row_idx - 1,
                student.user.name if student else 'N/A',
                student.user.email if student else 'N/A',
                student.cgpa if student else 0,
                student.user.department if student else 'N/A',
                guide.user.name if guide else 'N/A',
                alloc.status,
                alloc.method,
                f'Choice #{alloc.preference_rank}' if alloc.preference_rank > 0 else 'Fallback',
                alloc.allocated_at.strftime('%Y-%m-%d %H:%M') if alloc.allocated_at else 'N/A'
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='guide_allocation_report.xlsx'
        )
    except ImportError:
        flash('openpyxl is required for Excel export. Install it with: pip install openpyxl', 'error')
        return redirect(url_for('admin_dashboard'))


# ─── Notifications API ──────────────────────────────────────────
@app.route('/notifications/read/<int:notif_id>', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    notif = db.session.get(Notification, notif_id)
    if notif and notif.user_id == current_user.id:
        notif.read = True
        db.session.commit()
    return jsonify({'status': 'ok'})


@app.route('/notifications/read-all', methods=['POST'])
@login_required
def mark_all_read():
    Notification.query.filter_by(user_id=current_user.id, read=False).update({'read': True})
    db.session.commit()
    return jsonify({'status': 'ok'})


# ─── Context Processors ─────────────────────────────────────────
@app.context_processor
def inject_globals():
    unread_count = 0
    if current_user.is_authenticated:
        unread_count = Notification.query.filter_by(user_id=current_user.id, read=False).count()
    return {'unread_count': unread_count, 'now': datetime.now(timezone.utc)}


# ─── Error Handlers ─────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, message='Page not found'), 404


@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403, message='Access denied'), 403


# ─── Run ─────────────────────────────────────────────────────────
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)
